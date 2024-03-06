import pandas as pd
import openpyxl

##### fun(): LIN parameters #####
def lin_para():
    global version
    global speed
    global jitter_time
    global timebase_time
    while True:
        print('\n\n[LIN bus參數設置]')
        print(f'0: LIN protocol version: {version}')
        print(f'1: LIN speed: {speed} kbps')
        print(f'2: Jitter: {jitter_time} ms')
        print(f'3: Timebase: {timebase_time} ms')
        para_index = input("如需修改參數請輸入選項數字進行變更, 不需修改請按'Enter'繼續: ")

        if para_index == '':
            break
        else:
            match (para_index):
                case '0':
                    print(f'\n\nLIN protocol version 當前值為: {version}')
                    new_val = input("請輸入新的參數值, 如不需更改請按'Enter'回上頁: ")
                    if new_val != '':
                        version = new_val
                case '1':
                    print(f'\n\nLIN speed 當前值為: {speed} kbps')
                    new_val = input("請輸入新的參數值(輸入數字, 單位為kbps), 如不需更改請按'Enter'回上頁: ")
                    if new_val != '':
                        speed = new_val
                case '2':
                    print(f'\n\nJitter 當前值為: {jitter_time}')
                    new_val = input("請輸入新的參數值, 如不需更改請按'Enter'回上頁: ")
                    if new_val != '':
                        jitter_time = new_val
                case '3':
                    print(f'\n\nTimebase 當前值為: {timebase_time}')
                    new_val = input("請輸入新的參數值, 如不需更改請按'Enter'回上頁: ")
                    if new_val != '':
                        timebase_time = new_val
                case _:
                    pass

    
##### fun(): process data and generate data frame #####
# this function will remove Empty and Strikethrough format data
def process_data(workbook, sheet_name):
    worksheet = workbook[sheet_name]
    pData = []      # processed Data
    Sig_index = 6   # column index of 'Signal Name'
    Msg_index = 10  # column index of 'Message Name'
    Msb_index = 20  # column index of 'Msb'

    # for row_index in worksheet.iter_rows(values_only=True):
    for row_index in range(2, worksheet.max_row+1):
        Msg_value = worksheet.cell(row=row_index, column=Msg_index).value
        Msb_value = worksheet.cell(row=row_index, column=Msb_index).value
        Msg_strike = worksheet.cell(row=row_index, column=Msg_index).font.strike
        Sig_strike = worksheet.cell(row=row_index, column=Sig_index).font.strike

        # generate a processed Data
        if Msg_value and Msb_value != None and not Msg_strike and not Sig_strike:
            pData.append([worksheet.cell(row=row_index, column=col).value for col in range(1, worksheet.max_column + 1)])

    # get column name
    columns = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]

    # convert pData to DataFrame
    df = pd.DataFrame(pData, columns=columns)
    # convert 'Message ID' from Hex to Dec format
    df['Message ID'] = df['Message ID'].apply(lambda x: int(x, 16))
    return df


##### fun(): output_seg01 [LDF config] #####
def ldf_cfg(df):
    output_seg01 = '\n\nLIN_description_file;\n'
    output_seg01 += f'LIN_protocol_version = "{version}";\n'
    output_seg01 += f'LIN_language_version = "{version}";\n'
    output_seg01 += f'LIN_speed = {speed} kbps;\n'
    return output_seg01


##### fun(): output_seg02 [Nodes] #####
def ldf_notes(df):
    global slave_node_list
    notes_list = []
    Transmitter_list = []
    Receiver_list = []

    for index, row in df.iterrows():
        Transmitter = row['Transmitter']
        Receiver = row['Receiver']
        if Transmitter not in Transmitter_list:
            Transmitter_list.append(Transmitter)
        
        Receiver_split = Receiver.split('\n') # The Receiver may have mutiple rx-nodes
        for node in Receiver_split:
            if node not in Receiver_list:
                Receiver_list.append(node)
    notes_list = Transmitter_list + Receiver_list
    
    notes_list = list(set(notes_list)) # remove duplite data for notes_list
    notes_list.sort()
    print('\n\n[Node List]')
    for index, node in enumerate(notes_list):
        print(f'{index}: {node}')

    index = input("請輸入數字選擇一個Node作為Master node, 預設值:'0': ")
    if index == '':
        index = '0'
    master_node = notes_list[int(index)]

    slave_nodes = ''
    for node in notes_list:
        if node != master_node:
            slave_nodes += f'{node}, '
            slave_node_list.append(node) # for general used
    slave_nodes = slave_nodes[:-2]

    output_seg02 = '\nNodes {\n'
    output_seg02 += f'  Master: {master_node}, {timebase_time} ms, {jitter_time} ms ;\n'
    output_seg02 += f'  Slaves: {slave_nodes} ;\n'
    output_seg02 += '}\n'
    return output_seg02


##### fun(): output_seg03 [Signals] #####
def ldf_sig_def(df):
    output_seg03 = '\nSignals {\n'
    for index, row in df.iterrows():
        signal_name = row['Signal Name']
        size_bit = int(row['size(bit)'])
        init_val = row['Default Initialised value']
        if '0x' in str(init_val):
            init_val = int(row['Default Initialised value'], 16)
        tx_node = row['Transmitter']
        rx_node = row['Receiver']
        if '\n' in rx_node:
            rx_node = rx_node.replace('\n',', ')
        
        output_seg03 += f'  {signal_name}: {size_bit}, {init_val}, {tx_node}, {rx_node} ;\n'
    output_seg03 += '}\n'
    return output_seg03
        

##### fun(): output_seg04 [Diagnostic_signals] #####
def ldf_diag_sig(df):
    output_seg04 = '\nDiagnostic_signals {\n'
    output_seg04 += '  MasterReqB0: 8, 0 ;\n'
    output_seg04 += '  MasterReqB1: 8, 0 ;\n'
    output_seg04 += '  MasterReqB2: 8, 0 ;\n'
    output_seg04 += '  MasterReqB3: 8, 0 ;\n'
    output_seg04 += '  MasterReqB4: 8, 0 ;\n'
    output_seg04 += '  MasterReqB5: 8, 0 ;\n'
    output_seg04 += '  MasterReqB6: 8, 0 ;\n'
    output_seg04 += '  MasterReqB7: 8, 0 ;\n'
    output_seg04 += '  SlaveRespB0: 8, 0 ;\n'
    output_seg04 += '  SlaveRespB1: 8, 0 ;\n'
    output_seg04 += '  SlaveRespB2: 8, 0 ;\n'
    output_seg04 += '  SlaveRespB3: 8, 0 ;\n'
    output_seg04 += '  SlaveRespB4: 8, 0 ;\n'
    output_seg04 += '  SlaveRespB5: 8, 0 ;\n'
    output_seg04 += '  SlaveRespB6: 8, 0 ;\n'
    output_seg04 += '  SlaveRespB7: 8, 0 ;\n'
    output_seg04 += '}\n\n\n'
    return output_seg04


##### fun(): output_seg05 [Frames] #####
def ldf_data_frame_def(df):
    output_seg05 = '\nFrames {\n'
    df_group = df.groupby('Message Name')
    for group_index, group_data in df_group:
        frame_name = group_data['Message Name'].iloc[0]
        frame_id = group_data['Message ID'].iloc[0]
        tx_node = group_data['Transmitter'].iloc[0]
        frame_size = group_data['DLC'].iloc[0]
        output_seg05 += f'  {frame_name}: {frame_id}, {tx_node}, {frame_size} '
        output_seg05 += '{\n'
        for data_index in range(len(group_data)):
            signal_name = group_data['Signal Name'].iloc[data_index]
            start_bit = group_data['Lsb'].iloc[data_index]
            output_seg05 += f'    {signal_name}, {start_bit} ;\n'
        output_seg05 += '  }\n'
    output_seg05 += '}\n\n\n'
    return output_seg05


##### fun(): output_seg06 [Diagnostic_frames] #####
def ldf_diag_frame(df):
    output_seg06 = '\nDiagnostic_frames {\n'
    output_seg06 += '  MasterReq: 0x3c {\n'
    output_seg06 += '    MasterReqB0, 0 ;\n'
    output_seg06 += '    MasterReqB1, 8 ;\n'
    output_seg06 += '    MasterReqB2, 16 ;\n'
    output_seg06 += '    MasterReqB3, 24 ;\n'
    output_seg06 += '    MasterReqB4, 32 ;\n'
    output_seg06 += '    MasterReqB5, 40 ;\n'
    output_seg06 += '    MasterReqB6, 48 ;\n'
    output_seg06 += '    MasterReqB7, 56 ;\n'
    output_seg06 += '  }\n'
    output_seg06 += '  SlaveResp: 0x3d {\n'
    output_seg06 += '    SlaveRespB0, 0 ;\n'
    output_seg06 += '    SlaveRespB1, 8 ;\n'
    output_seg06 += '    SlaveRespB2, 16 ;\n'
    output_seg06 += '    SlaveRespB3, 24 ;\n'
    output_seg06 += '    SlaveRespB4, 32 ;\n'
    output_seg06 += '    SlaveRespB5, 40 ;\n'
    output_seg06 += '    SlaveRespB6, 48 ;\n'
    output_seg06 += '    SlaveRespB7, 56 ;\n'
    output_seg06 += '  }\n'
    output_seg06 += '}\n'
    return output_seg06


##### fun(): output_seg07 [Node_attributes] #####
def ldf_node_attr(df):
    global slave_node_list
    config_nad = 1 # 1-255 (Hex/Dec)
    init_nad = 1
    supplier_id = '0x0' # 0-0x7FFE
    finction_id = '0x0' # 0-0xFFFE
    variant = '0x0' # 0-0xFF
    output_seg07 = '\nNode_attributes {\n'
    for slave_node in slave_node_list:
        output_seg07 += f'  {slave_node}'
        output_seg07 += '{\n'
        output_seg07 += f'    LIN_protocol = "{version}" ;\n'
        output_seg07 += f'    configured_NAD = {config_nad} ;\n'
        output_seg07 += f'    initial_NAD = {init_nad} ;\n'
        output_seg07 += f'    product_id = {supplier_id}, {finction_id}, {variant} ;\n' # 
        output_seg07 += '    P2_min = 50 ms ;\n'
        output_seg07 += '    ST_min = 0 ms ;\n'
        output_seg07 += '    N_As_timeout = 1000 ms ;\n'
        output_seg07 += '    N_Cr_timeout = 1000 ms ;\n'
        output_seg07 += '    configurable_frames {\n'
        # change NAD for next salve node, NAD should be unique
        config_nad = config_nad + 1
        init_nad = init_nad + 1

        df_group = df.groupby('Message Name')
        for group_index, group_data in df_group:
            frame = group_data['Message Name'].iloc[0]
            tx_node = group_data['Transmitter'].iloc[0]
            rx_node_list = group_data['Receiver'].iloc[0].split('\n')
            # check whether the slave node is related to the Transmitter or Receiver of the message
            # if so, the message belongs configurable_frames
            if slave_node == tx_node or slave_node in rx_node_list:
                output_seg07 += f'      {frame} ;\n'

        output_seg07 += '    }\n'
        output_seg07 += '  }\n'
    output_seg07 += '}\n'
    return output_seg07


##### fun(): output_seg08 [Schedule_tables] #####
def ldf_sch_table(excel_file):
    # load excel without column name (header=None)
    df = pd.read_excel(excel_file, sheet_name='LIN_Schedule Table', header=None)

    # add new column name
    df.columns = ['No', 'Time', 'Delay', 'Message', 'Cycle', 'Comment']
   
    # search schedule table in sheet 'LIN_Schedule Table'
    table_list = []
    for index, row in df.iterrows():
        text = str(row.iloc[0]).lower()
        if 'schedule table' in text or 'schedule_table' in text:
            table_list.append(row.iloc[0])
            #print(row.iloc[0])
    
    print('\n\n[Schedule table list]')
    for index, tableName in enumerate(table_list):
        print(f'{index}: {tableName}')
    table_index = input("請輸入數字選擇一個Schedule Table (若Schedule Table不正確會導致LDF不可用), 預設值為'0': ")
    if table_index == '':
        table_index = '0'
    selected_table = table_list[int(table_index)]

    #selected_table = 'CEM_LIN1 Schedule Table（IG_ON/IG_OFF）'
    
    # dectect the starting row and endinf row of data
    table_found = False
    # due to the last row is not empty, add a empty row at button of the df for end_index
    df.loc[len(df)] = pd.Series()

    for index, row in df.iterrows():
        data = row['Message']
        if selected_table in str(row.iloc[0]):
            start_index = index+2
            table_found = True
        if table_found and index > start_index and pd.isna(data):
            end_index = index
            break

    output_seg08 = '\nSchedule_tables {\n'
    output_seg08 += ' Table1 {\n'
    for index, row in df.iloc[start_index:end_index].iterrows():
        frame_name = df['Message'].iloc[index]
        delay_time = df['Delay'].iloc[index]
        output_seg08 += f'    {frame_name} delay {delay_time} ms ;\n'
    output_seg08 += '  }\n'
    output_seg08 += '}\n'
    return output_seg08
            

##### fun(): output_seg09 [Signal_encoding_types] #####
def ldf_sig_encode(df):
    output_seg09 = '\nSignal_encoding_types {\n'
    for index, row in df.iterrows():
        signal_name = row['Signal Name']
        minimum = row['P-Minimum']
        maxmum = row['P-Maximum']
        factor = row['Factor']
        offect = row['Offset']
        unit = row['Unit']
        coding = row['Coding']

        if not pd.isna(signal_name):
            sig_encoding = 'Enc_' + signal_name
            output_seg09 += f'  {sig_encoding} '
            output_seg09 += '{\n'
            output_seg09 += f'    physical_value, {minimum}, {maxmum}, {factor}, {offect}, "{unit}" ;\n'
            
            if not pd.isna(coding):
                coding = coding.replace('=',':')
                coding = coding.replace(' : ',':')
                coding = coding.replace(' :',':')
                coding = coding.replace(': ',':')
                coding = coding.replace(' \n','\n')
                coding_list = coding.split('\n')
                for coding_data in coding_list:
                    if '~' not in coding_data:
                        coding_data_list = coding_data.split(':')
                        value = coding_data_list[0]
                        value = int(value, 16)
                        description = coding_data_list[1]
                        output_seg09 += f'    logical_value, {value}, "{description}" ;\n'
            output_seg09 += '  }\n'
    output_seg09 += '}\n'
    return output_seg09


##### fun(): output_seg10 [Signal_representation] #####
def ldf_sig_represent(df):
    output_seg10 = '\nSignal_representation {\n'
    for index, row in df.iterrows():
        signal_name = row['Signal Name']
        if not pd.isna(signal_name):
            output_seg10 += f'  Enc_{signal_name}: {signal_name} ;\n'
    output_seg10 += '}\n'
    return output_seg10



##### Main #####
def ldf_main():
    while True:
        try:
            # load Excel file
            print('\nThis program will generate a LDF file from excel.\n')
            excel_file = input("請輸入欲轉換的Excel檔名: ")
            if not '.xlsx' in excel_file:
                excel_file += '.xlsx'
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            break
        except:
            print('\nError! 檔名錯誤或找不到檔案路徑')
            input('Press [Enter] to continue.\n')

    while True:
        # get sheetName and sheet_index from workbook
        sheet_name_list = workbook.sheetnames
        print('\n\n[Sheet list]')
        for index, sheetName in enumerate(sheet_name_list):
            print(f'{index}: {sheetName}')

        sheet_index = input("請輸入數字選擇一個sheet生成LDF, 或輸入'q'結束程式: ")
        
        if sheet_index.lower() == 'q':
            break
        else:
            try:
                sheetName = sheet_name_list[int(sheet_index)]
                # process data and generate df(data frame)
                df = process_data(workbook, sheetName)
                
                # declare variable for LIN bus config
                global version
                global speed
                global jitter_time
                global timebase_time
                version = '2.1'
                speed = '19.2'
                jitter_time = '0.1'
                timebase_time ='10'
                # declare slave_node_list for ldf_notes() and ldf_node_attr()
                global slave_node_list
                slave_node_list = []

                # setup LIN bus parameter
                lin_para()

                output_01 = ldf_cfg(df)
                output_02 = ldf_notes(df)
                output_03 = ldf_sig_def(df)
                output_04 = ldf_diag_sig(df)
                output_05 = ldf_data_frame_def(df)
                output_06 = ldf_diag_frame(df)
                output_07 = ldf_node_attr(df)     
                output_08 = ldf_sch_table(excel_file)
                output_09 = ldf_sig_encode(df)
                output_10 = ldf_sig_represent(df)
                output_text = output_01 + output_02 + output_03 + output_04 + output_05 + output_06 + output_07 + output_08 + output_09 + output_10

                LDF_name = sheetName + '.ldf'
                with open(LDF_name, 'w', encoding='utf-8') as f:
                    f.write(output_text)
                print(f'\n{LDF_name} is generated!!\n')
                input('Press [Enter] to continue.')
            except:
                print('\nError! 請確認所選擇的sheet內容是否正確')
                input('Press [Enter] to continue.\n')
        
        '''sheetName = sheet_name_list[int(sheet_index)]
        # process data and generate df(data frame)
        df = process_data(workbook, sheetName)
        
        # declare variable for LIN bus config
        global version
        global speed
        global jitter_time
        global timebase_time
        version = '2.1'
        speed = '19.2'
        jitter_time = '0.1'
        timebase_time ='10'
        # declare slave_node_list for ldf_notes() and ldf_node_attr()
        global slave_node_list
        slave_node_list = []

        # setup LIN bus parameter
        lin_para()

        output_01 = ldf_cfg(df)
        output_02 = ldf_notes(df)
        output_03 = ldf_sig_def(df)
        output_04 = ldf_diag_sig(df)
        output_05 = ldf_data_frame_def(df)
        output_06 = ldf_diag_frame(df)
        output_07 = ldf_node_attr(df)     
        output_08 = ldf_sch_table(excel_file)
        output_09 = ldf_sig_encode(df)
        output_10 = ldf_sig_represent(df)
        output_text = output_01 + output_02 + output_03 + output_04 + output_05 + output_06 + output_07 + output_08 + output_09 + output_10
        LDF_name = sheetName + '.ldf'
        with open(LDF_name, 'w', encoding='utf-8') as f:
            f.write(output_text)
        print('\nLDF is generated!!\n')
        #print(slave_node_list)
        input('Press [Enter] to continue.')'''

if __name__ == "__main__":
    ldf_main()