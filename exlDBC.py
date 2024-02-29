import pandas as pd
import openpyxl
#import re

##### fun(): check for special cases in signal names #####
def chk_signalname(sigName):
    sigName = sigName.replace('(PS:自定义)','')
    sigName = sigName.replace(' \n(PS: 自定义)','')
    sigName = sigName.replace(' ','')
    if '\n' in sigName and not 'EMMC_BYTE_' in sigName:
        sigName_split = sigName.split('\n')
        final_sigName = sigName_split[len(sigName_split)-1]
        #print(sigName,'  --->  ', final_sigName)
    else:
        sigName = sigName.replace('\n','')
        final_sigName = sigName
    return final_sigName

##### fun(): process val for SG_ #####
def sg_value(group_data, colName, dataIndex):
    val = group_data[colName].iloc[dataIndex]
    val = str(val).replace("\n(0xFF)","")   # remove "\n(0xFF)"
    # convert empty data to 0
    if pd.isna(val) or val == 'nan' or val == 'None':
        if colName == 'Factor':
            val = 1
        else:
            val = 0
    else:
        pass
    return val


##### fun(): process data and generate data frame #####
# this function will remove Empty and Strikethrough format data
def process_data(workbook, sheet_name):
    worksheet = workbook[sheet_name]
    pData = []      # processed Data
    Sig_index = 6   # column index of 'Signal Name'
    Msg_index = 10  # column index of 'Message Name'
    Mab_index = 20  # column index of 'Mab'

    # for row_index in worksheet.iter_rows(values_only=True):
    for row_index in range(2, worksheet.max_row+1):
        Msg_value = worksheet.cell(row=row_index, column=Msg_index).value
        Mab_value = worksheet.cell(row=row_index, column=Mab_index).value
        Msg_strike = worksheet.cell(row=row_index, column=Msg_index).font.strike
        Sig_strike = worksheet.cell(row=row_index, column=Sig_index).font.strike

        if Msg_value and Mab_value and not Msg_strike and not Sig_strike:
            # generate a processed Data
            pData.append([worksheet.cell(row=row_index, column=col).value for col in range(1, worksheet.max_column + 1)])

    # get column name
    columns = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]

    # convert pData to DataFrame
    df = pd.DataFrame(pData, columns=columns)
    # convert 'Message ID' from Hex to Dec format
    df['Message ID'] = df['Message ID'].apply(lambda x: int(x, 16))
    return df

##### fun(): output_seg01 [VERSION | NS_ | BS_] #####
def dbc_ver_ns_bs():
    # VERSION
    output_seg01 = 'VERSION ""\n\n\n'

    # NS_
    output_seg01 += 'NS_ :\n'
    output_seg01 += '    NS_DESC_\n    CM_\n    BA_DEF_\n    BA_\n    VAL_\n'
    output_seg01 += '	CAT_DEF_\n    CAT_\n	FILTER\n	BA_DEF_DEF_\n    EV_DATA_\n'
    output_seg01 += '    ENVVAR_DATA_\n    SGTYPE_\n    SGTYPE_VAL_\n    BA_DEF_SGTYPE_\n'
    output_seg01 += '    BA_SGTYPE_\n    SIG_TYPE_REF_\n    VAL_TABLE_\n    SIG_GROUP_\n'
    output_seg01 += '    SIG_VALTYPE_\n    SIGTYPE_VALTYPE_\n    BO_TX_BU_\n    BA_DEF_REL_\n'
    output_seg01 += '    BA_REL_\n    BA_DEF_DEF_REL_\n    BU_SG_REL_\n    BU_EV_REL_\n'
    output_seg01 += '    BU_BO_REL_\n    SG_MUL_VAL_\n\n'

    # BS_
    output_seg01 += 'BS_:\n\n'
    return output_seg01


##### fun(): output_seg02 [BU_] #####
def dbc_bu(df):
    # BU_: {node_name_1} {node_name_2} ...
    output_seg02 = 'BU_: '

    # add node_name from 'Transmitter'
    df_group = df.groupby('Transmitter')
    Transmitter = list(df_group.groups.keys())
    for node_name in Transmitter:
        output_seg02 += f'{node_name} '

    # add node_name from 'Receiver' but exclude duplicates in 'Transmitter'
    df_group = df.groupby('Receiver')
    Receiver = list(df_group.groups.keys())
    for node_name in Receiver:
        node_name = node_name.replace("\n","/")
        node_name_split = node_name.split("/")
        for receiver_node_name in node_name_split:
            if not receiver_node_name in Transmitter:
                output_seg02 += f'{receiver_node_name} '               

    output_seg02 += '\n\n\n'
    return output_seg02


##### fun(): output_seg03 [BO_ | SG_] #####
def dbc_bo_sg(df):
    # BO_ 
    output_seg03 =""
    df_group = df.groupby('Message Name')
    # get the first data of each group through group_index
    for group_index, group_data in df_group:
        transmitter = group_data['Transmitter'].iloc[0]
        if pd.isna(transmitter):
            transmitter = 'Vector__XXX'
        message_name = group_data['Message Name'].iloc[0]
        message_id = group_data['Message ID'].iloc[0]
        message_size = group_data['DLC'].iloc[0]
        # BO_ {message_id} {message_name}: {message_size} {transmitter}
        output_seg03 += f'BO_ {message_id} {message_name}: {message_size} {transmitter}\n'

    # SG_
        for dataIndex in range(len(group_data)):
            # signal_name
            signal_name = group_data['Signal Name'].iloc[dataIndex]
            signal_name = chk_signalname(signal_name)

            # multiplexer_indicator
            multiplexer_indicator = ''

            # signal_size
            signal_size = group_data['size(bit)'].iloc[dataIndex].astype(int)

            # factor
            factor = sg_value(group_data, 'Factor', dataIndex)

            # offset
            offset = sg_value(group_data, 'Offset', dataIndex)

            # minimum
            minimum = sg_value(group_data, 'P-Minimum', dataIndex)

            # maximum
            maximum = sg_value(group_data, 'P-Maximum', dataIndex)

            # start_bit | byte_order
            ByteOrder = group_data['Byte Order'].iloc[dataIndex]
            if ByteOrder == 'Motorola':
                start_bit = group_data['Mab'].iloc[dataIndex].astype(int)
                byte_order = '0'
            else:   # Intel
                start_bit = group_data['Lab'].iloc[dataIndex].astype(int)
                byte_order = '1'
            
            # value_type
            DataType = group_data['Data Type'].iloc[dataIndex]
            if DataType == 'unsigned':
                value_type = '+'
            else:   # signed
                value_type = '-'
            
            # unit
            unit = group_data['Unit'].iloc[dataIndex]
            if pd.isna(unit):
                unit = ''
            
            # receiver
            receiver = group_data['Receiver'].iloc[dataIndex]
            receiver = receiver.replace('\n',',')
            receiver = receiver.replace('/',',')

            # SG_ {signal_name} {multiplexer_indicator} : {start_bit}|{signal_size}@{byte_order}{value_type} ({factor},{offset}) [{minimum}|{maximum}] "{unit}" {receiver}
            if not 'EMMC_BYTE_' in signal_name:
                output_seg03 += f' SG_ {signal_name} {multiplexer_indicator}: {start_bit}|{signal_size}@{byte_order}{value_type} ({factor},{offset}) [{minimum}|{maximum}] "{unit}" {receiver}\n'
            # only for EMMC_BYTE_# signals
            else:
                signal_name = signal_name.replace('EMMC_BYTE_','')
                signal_name_split = signal_name.split("~")
                start_bit = 7
                for number in range(int(signal_name_split[0]), int(signal_name_split[1])+1):
                    signal_name = f'EMMC_BYTE_{number}'
                    signal_size = 8
                    output_seg03 += f' SG_ {signal_name} {multiplexer_indicator}: {start_bit}|{signal_size}@{byte_order}{value_type} ({factor},{offset}) [{minimum}|{maximum}] "{unit}" {receiver}\n'
                    start_bit = start_bit+8

            if dataIndex == len(group_data)-1:
                output_seg03 += '\n'
    output_seg03 += '\n\n'
    return output_seg03


##### fun(): output_seg04 [BA_DEF_] #####
def dbc_ba_def():
    global ILSupport
    global busType
    # BA_DEF_
    output_seg04 = ''
    output_seg04 += 'BA_DEF_  "BusType" STRING ;\n'
    output_seg04 += 'BA_DEF_ BO_  "GenMsgFastOnStart" INT 0 0;\n'
    output_seg04 += 'BA_DEF_ SG_  "GenSigInactiveValue" INT 0 0;\n'
    output_seg04 += 'BA_DEF_ BU_  "ILUsed" ENUM  "Yes","No";\n'
    output_seg04 += 'BA_DEF_ SG_  "GenSigStartValue" FLOAT 0 100000000000;\n'
    output_seg04 += 'BA_DEF_ SG_  "GenSigSendType" ENUM  "Cyclic","OnWrite","OnWriteWithRepetition","OnChange","OnChangeWithRepetition","IfActive","IfActiveWithRepetition","NoSigSendType","OnChangeAndIfActive","OnChangeAndIfActiveWithRepetition";\n'
    output_seg04 += 'BA_DEF_ BO_  "GenMsgNrOfRepetition" INT 0 999999;\n'
    output_seg04 += 'BA_DEF_ BO_  "GenMsgDelayTime" INT 0 1000;\n'
    output_seg04 += 'BA_DEF_ BO_  "GenMsgCycleTime" INT 0 50000;\n'
    output_seg04 += 'BA_DEF_ BO_  "GenMsgSendType" ENUM  "Cyclic","not_used","not_used","not_used","not_used","not_used","not_used","IfActive","NoMsgSendType";\n'
    output_seg04 += 'BA_DEF_ BO_  "GenMsgCycleTimeFast" INT 0 100000;\n'
    output_seg04 += 'BA_DEF_ BO_  "GenMsgILSupport" ENUM  "No","Yes";\n'
    output_seg04 += 'BA_DEF_ BO_  "GenMsgStartDelayTime" INT 0 100000;\n'
    output_seg04 += 'BA_DEF_ BO_  "VFrameFormat" ENUM  "StandardCAN","ExtendedCAN","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","reserved","StandardCAN_FD","ExtendedCAN_FD";\n'
    output_seg04 += 'BA_DEF_ BU_  "NodeLayerModules" STRING ;\n'
    # BA_DEF_DEF_
    output_seg04 += 'BA_DEF_DEF_  "BusType" "";\n'
    output_seg04 += 'BA_DEF_DEF_  "GenMsgFastOnStart" 0;\n'
    output_seg04 += 'BA_DEF_DEF_  "GenSigInactiveValue" 0;\n'
    output_seg04 += 'BA_DEF_DEF_  "ILUsed" "Yes";\n'
    output_seg04 += 'BA_DEF_DEF_  "GenSigStartValue" 0;\n'
    output_seg04 += 'BA_DEF_DEF_  "GenSigSendType" "Cyclic";\n'
    output_seg04 += 'BA_DEF_DEF_  "GenMsgNrOfRepetition" 0;\n'
    output_seg04 += 'BA_DEF_DEF_  "GenMsgDelayTime" 0;\n'
    output_seg04 += 'BA_DEF_DEF_  "GenMsgCycleTime" 100;\n'
    output_seg04 += 'BA_DEF_DEF_  "GenMsgSendType" "Cyclic";\n'
    output_seg04 += 'BA_DEF_DEF_  "GenMsgCycleTimeFast" 0;\n'
    output_seg04 += f'BA_DEF_DEF_  "GenMsgILSupport" "{ILSupport}";\n'
    output_seg04 += 'BA_DEF_DEF_  "GenMsgStartDelayTime" 0;\n'
    output_seg04 += 'BA_DEF_DEF_  "VFrameFormat" "StandardCAN";\n'
    output_seg04 += 'BA_DEF_DEF_  "NodeLayerModules" "CANoeILNLVector.dll";\n'
    output_seg04 += f'BA_ "BusType" "{busType}";\n'
    return output_seg04

##### fun(): output_seg05 [BA_] & output_seg06 [VAL_] #####
def dbc_ba(df):
    # BA_
    output_seg05 = ''
    
    # get the first data of each group through group_index
    df_group = df.groupby('Message Name')
    for group_index, group_data in df_group:
        message_id = group_data['Message ID'].iloc[0]
        message_type = group_data['Message Type'].iloc[0]
        message_size = group_data['DLC'].iloc[0].astype(int)

        # identify MsgSendType
        if message_type == 'P':
            MsgCycleTime = group_data['period\n(ms)'].iloc[0].astype(int)
            output_seg05 += f'BA_ "GenMsgCycleTime" BO_ {message_id} {MsgCycleTime};\n'
        elif message_type == 'E' or message_type == 'M':
            # setup for "GenMsgSendType"
            MsgSendType = '8'   # NoMsgSendType
            output_seg05 += f'BA_ "GenMsgSendType" BO_ {message_id} {MsgSendType};\n'
        else:
            pass

        # identify CANFD
        if message_size > 8:
            VFrameFormat = '14'   #StandardCAN_FD
            output_seg05 += f'BA_ "VFrameFormat" BO_ {message_id} {VFrameFormat};\n'
        
            for dataIndex in range(len(group_data)):
                signal_name = group_data['Signal Name'].iloc[dataIndex]
                signal_name = chk_signalname(signal_name)
                SigSendType = 1 # OnWrite
                if not 'EMMC_BYTE_' in signal_name:
                    output_seg05 += f'BA_ "GenSigSendType" SG_ {message_id} {signal_name} {SigSendType};\n'
    return output_seg05


##### fun(): output_seg06 [VAL_] #####
def dbc_val(df):
    # VAL_
    output_seg06 = ''
    df_group = df.groupby('Message Name')
    for group_index, group_data in df_group:
        message_id = group_data['Message ID'].iloc[0]

        for dataIndex in range(len(group_data)):
            # A flag used to determine whether to generate a ValTable
            gen_ValTable = True

            ValTable = group_data['Coding'].iloc[dataIndex]
            # if ValTable is not empty, create a value table for the signal
            if not pd.isna(ValTable):
                # process the contents of ValTable
                ValTable = ValTable.replace(' : ',':')
                ValTable = ValTable.replace('：',':')
                ValTable = ValTable.replace(': ',':')
                ValTable = ValTable.replace(' :',':')
                ValTable = ValTable.replace(' ~ ','~')
                ValTable = ValTable.replace(':~','~')
                ValTable = ValTable.replace('-0','~0')
                ValTable = ValTable.replace('Ox','0x')
                ValTable = ValTable.replace('"','')
                ValTable = ValTable.replace('0x16-1F','0x16~1F')
                ValTable = ValTable.replace(' 0x','\n0x')   # SAS_Sts (CAN01_Matrix\INT3)

                # process the signal_name
                signal_name = group_data['Signal Name'].iloc[dataIndex]
                signal_name = chk_signalname(signal_name)
                
                # split the ValTable by '\n' and save the results to a list
                ValTable_split = ValTable.split("\n")

                Value_Description = ''
                for data in ValTable_split:
                    r"""
                    if '~' in data:     
                        # use regular expression to separate data (0xAA~0xBB:CCCC) into three parts:'0xAA', '0xBB' and 'CCCC'
                        pattern = re.compile(r'(\w+x\w+)~(\w+x\w+):(\w+)')
                        regular_data = pattern.match(data)
                        start, end, Description = regular_data.groups()
                        for Value in range(int(start, 16), int(end, 16)+1):
                            Value_Description = f'{Value} "{Description}" '
                            output_seg06 += f'VAL_ {message_id} {signal_name} {Value_Description};\n'
                        
                        # do not porcess data containing '~', because the range of some value is too large
                        gen_ValTable = False
                        continue
                    elif 'EMMC_BYTE_' in signal_name:
                    """

                    if 'EMMC_BYTE_' in signal_name:
                        # ValTable is not generated when the signal name contains "EMMC_BYTE_"
                        gen_ValTable = False
                    else:
                        # split data to two elements (Value & Description) by ':'
                        data_split = data.split(":", 1)

                        # due to some data lacks ':', use " " to split instead
                        if len(data_split) == 1:
                            data_split = data.split(" ", 1)
                        
                        # do not porcess data containing '~'
                        if '~' in data_split[0]:
                            Value =''
                            Description =''
                        else:
                            Value = int(data_split[0], 16)  # convert Hex to Dec
                            Description = f'"{data_split[1]}" '
                        Value_Description += f'{Value} {Description}'
                
                # if gen_ValTable is True, generates ValTable
                if gen_ValTable:
                    output_seg06 += f'VAL_ {message_id} {signal_name} {Value_Description};\n'
    return output_seg06

    
##### Main #####
def dbc_main():
    while True:
        try:
            # load Excel file
            print('\nThis program will generate a DBC file from excel.\n')
            excel_file = input('請輸入欲轉換的Excel檔名: ')
            if not '.xlsx' in excel_file:
                excel_file += '.xlsx'
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            break
        except:
            print('\nError! 檔名錯誤或找不到檔案路徑')
            input('Press [Enter] to continue.\n')

    while True:
        # get sheetName and sheet_index from workbook
        sheet_name_list = workbook.sheetnames
        
        print('\n\n[Sheet list]')
        for index, sheetName in enumerate(sheet_name_list):
            print(f'{index}: {sheetName}')
        sheet_index = input("請輸入數字選擇一個sheet生成DBC, 或輸入'q'結束程式: ")
        if sheet_index.lower() == 'q':
            break

        # determine the BusType attribute
        else:
            global busType
            global ILSupport

            print('\n\n[CAN bus types]\n0. CAN\n1. CAN FD')
            busType = input("請輸入數字選擇CAN bus類別, 或輸入'q'結束程式, 預設值為'0': ")
            if busType == '1':
                busType = 'CAN FD'
            elif busType.lower() == 'q':
                break
            else:
                busType = 'CAN'

            # Determine the GenMsgILSupport attribute
            print('\n\n[GenMsgILSupport]\n0. No\n1. Yes')
            print('(若啟用GenMsgILSupport, CANoe會根據Attribute定義自動發送CAN message.)')
            ILSupport = input("請輸入數字選擇是否啟用GenMsgILSupport, 或輸入'q'結束程式, 預設值為'0':")
            if ILSupport == '1':
                ILSupport = 'Yes'
            elif ILSupport.lower() == 'q':
                break
            else:
                ILSupport = 'No'

            # generate DBC files          
            try:
                sheetName = sheet_name_list[int(sheet_index)]
                df = process_data(workbook, sheetName)

                output_01 = dbc_ver_ns_bs()
                output_02 = dbc_bu(df)
                output_03 = dbc_bo_sg(df)
                output_04 = dbc_ba_def()
                output_05 = dbc_ba(df)
                output_06 = dbc_val(df)
                
                # create a DBC file
                output_text = output_01 + output_02 + output_03 + output_04 + output_05 + output_06
                DBC_name = sheetName+'.dbc'
                with open(DBC_name, 'w', encoding='utf-8') as f:
                    f.write(output_text)
                print(f'\n{DBC_name} is generated!!\n')
                input('Press [Enter] to continue.')
            except:
                print('\nError! 請確認所選擇的sheet內容是否正確')
                input('Press [Enter] to continue.\n')
            
if __name__ == '__main__':
    dbc_main()